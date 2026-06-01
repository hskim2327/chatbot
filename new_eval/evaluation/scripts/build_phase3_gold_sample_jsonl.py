"""Phase 3 RFP 도메인 정답지 산출물을 생성한다.

이 스크립트는 사용자가 승인한 hybrid_50 추천안을 최종 Phase 3 pilot
gold set으로 변환한다. 평가 metric을 구현하지 않으며, 원본 RFP 전문이나
긴 표 원문을 산출물에 저장하지 않는다.
"""

from __future__ import annotations

import argparse
import ast
import csv
import difflib
import json
import math
import platform
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BATCH_VERIFICATION_METHOD = "batch_user_approval_from_hybrid_50"
BATCH_REVIEWER = "project_owner_batch_approval"
BATCH_REVIEW_NOTES = (
    "User approved hybrid_50 as Phase 3 pilot gold set without row-level Excel edits."
)
ALLOWED_REFUSAL_PHRASES = [
    "확인할 수 없습니다",
    "명시되어 있지 않습니다",
    "제공된 자료에는 없습니다",
]
DEFAULT_FORBIDDEN_CLAIM_TYPES = [
    "문서에 없는 정보 단정",
    "계약 결과 단정",
    "낙찰업체 단정",
    "내부 의도 단정",
]
DEFAULT_OUTPUT_STRUCTURE = ["문서별 요약", "공통점", "차이점"]
TASK_LABELS = {
    "budget": "예산/금액 평가",
    "multi_doc_comparison": "여러 문서 비교 평가",
    "required_fields": "필수 정보 평가",
    "submission_eligibility_deadline": "제출서류/입찰자격/마감일 평가",
    "unanswerable": "문서에 없는 질문 대응 평가",
    "robust_query_type_e": "오타/구어체 견고성 평가",
}
TASK_BLOCK_NAMES = {
    "budget": "budget_gold",
    "multi_doc_comparison": "multi_doc_comparison_gold",
    "required_fields": "required_field_gold",
    "submission_eligibility_deadline": "submission_eligibility_deadline_gold",
    "unanswerable": "unanswerable_gold",
    "robust_query_type_e": "robust_query_gold",
}
SECRET_PATTERNS = [
    re.compile(r"sk-[A-Za-z0-9_\-]{20,}"),
    re.compile(r"AKIA[0-9A-Z]{16}"),
    re.compile(r"-----BEGIN [A-Z ]*PRIVATE KEY-----"),
    re.compile(r"api[_-]?key\s*[:=]", re.IGNORECASE),
]
MAX_CELL_TEXT = 500
MAX_LIST_ITEM_TEXT = 180
MAX_EVIDENCE_TEXT = 180


@dataclass
class GenerationResult:
    """정답지 생성 결과를 한 번에 전달하기 위한 묶음이다."""

    records: list[dict[str, Any]]
    readable_rows: list[dict[str, Any]]
    validation_rows: list[dict[str, Any]]
    approved_df: pd.DataFrame


def parse_args() -> argparse.Namespace:
    """CLI 인자를 정의한다."""

    project_root = Path.cwd()
    base = project_root / "eval" / "evaluation"
    data_dir = base / "data"
    docs_dir = base / "docs"
    scripts_dir = base / "scripts"

    parser = argparse.ArgumentParser(
        description="hybrid_50 기반 Phase 3 최종 gold JSONL/XLSX/CSV를 생성한다."
    )
    parser.add_argument("--project-root", type=Path, default=project_root)
    parser.add_argument(
        "--input-hybrid-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_hybrid_50_recommendation.csv",
    )
    parser.add_argument(
        "--review-xlsx",
        type=Path,
        default=data_dir / "rfp_domain_gold_review.xlsx",
    )
    parser.add_argument(
        "--output-jsonl",
        type=Path,
        default=data_dir / "rfp_domain_gold_sample.jsonl",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=data_dir / "rfp_domain_gold_sample.xlsx",
    )
    parser.add_argument(
        "--output-readable-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_sample_readable.csv",
    )
    parser.add_argument(
        "--output-approved-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_hybrid_50_approved.csv",
    )
    parser.add_argument(
        "--output-validation-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_sample_validation.csv",
    )
    parser.add_argument(
        "--output-guide",
        type=Path,
        default=docs_dir / "rfp_domain_gold_sample_guide.md",
    )
    parser.add_argument(
        "--output-report",
        type=Path,
        default=docs_dir / "phase3_gold_jsonl_generation_report.md",
    )
    parser.add_argument(
        "--output-cleanup-manifest",
        type=Path,
        default=docs_dir / "phase3_cleanup_manifest.md",
    )
    parser.add_argument(
        "--selected-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_selected_50.csv",
    )
    parser.add_argument(
        "--extension-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_phase3_extension_candidates.csv",
    )
    parser.add_argument(
        "--candidates-csv",
        type=Path,
        default=data_dir / "rfp_domain_gold_candidates.csv",
    )
    parser.add_argument(
        "--scripts-dir",
        type=Path,
        default=scripts_dir,
    )
    parser.add_argument("--cleanup", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--resolve-warnings", action="store_true")
    parser.add_argument(
        "--chunks-jsonl",
        type=Path,
        default=project_root / "new_data" / "chunks_v2_690.jsonl",
    )
    parser.add_argument(
        "--eval-dir",
        type=Path,
        default=project_root / "data" / "eval",
    )
    parser.add_argument(
        "--output-warning-resolution-report",
        type=Path,
        default=docs_dir / "phase3_gold_warning_resolution_report.md",
    )
    return parser.parse_args()


def is_blank(value: Any) -> bool:
    """빈 값 여부를 판정한다."""

    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if isinstance(value, str) and value.strip().lower() in {"nan", "none", "null"}:
        return True
    return False


def clean_text(value: Any, max_len: int = MAX_LIST_ITEM_TEXT) -> str:
    """긴 원문 저장을 피하기 위해 텍스트를 짧게 정리한다."""

    if is_blank(value):
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_len:
        return text[: max_len - 14].rstrip() + " ...[truncated]"
    return text


def safe_json_loads(value: Any) -> Any:
    """JSON 문자열 또는 Python literal 형태의 값을 안전하게 파싱한다."""

    if is_blank(value):
        return []
    if isinstance(value, (list, dict)):
        return value
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if text == "":
        return []
    for loader in (json.loads, ast.literal_eval):
        try:
            return loader(text)
        except Exception:
            pass
    return text


def as_list(value: Any) -> list[Any]:
    """값을 리스트로 정규화한다."""

    parsed = safe_json_loads(value)
    if is_blank(parsed):
        return []
    if isinstance(parsed, list):
        return parsed
    if isinstance(parsed, tuple):
        return list(parsed)
    if isinstance(parsed, dict):
        return [parsed]
    return [parsed]


def as_clean_list(value: Any, max_len: int = MAX_LIST_ITEM_TEXT) -> list[str]:
    """문자 리스트를 짧은 문자열 리스트로 정규화한다."""

    result: list[str] = []
    for item in as_list(value):
        if isinstance(item, dict):
            result.append(clean_text(json.dumps(item, ensure_ascii=False), max_len))
        else:
            text = clean_text(item, max_len)
            if text:
                result.append(text)
    return dedupe(result)


def dedupe(values: list[Any]) -> list[Any]:
    """순서를 유지하며 중복을 제거한다."""

    seen: set[str] = set()
    result: list[Any] = []
    for value in values:
        key = json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list)) else str(value)
        if key not in seen:
            seen.add(key)
            result.append(value)
    return result


def json_dumps(value: Any) -> str:
    """CSV/Excel에 넣을 복합 값을 JSON 문자열로 변환한다."""

    return json.dumps(value, ensure_ascii=False)


def to_builtin(value: Any) -> Any:
    """pandas/numpy 값을 JSON 직렬화 가능한 Python 기본 타입으로 바꾼다."""

    if isinstance(value, dict):
        return {key: to_builtin(item) for key, item in value.items()}
    if isinstance(value, list):
        return [to_builtin(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value
    return value


def read_csv_if_exists(path: Path) -> pd.DataFrame:
    """CSV가 있으면 읽고, 없으면 빈 DataFrame을 반환한다."""

    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, encoding="utf-8-sig")


def read_workbook_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    """검수 workbook sheet를 읽는다. 없으면 빈 DataFrame을 반환한다."""

    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def row_by_id(df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    """id를 key로 하는 행 사전을 만든다."""

    if df.empty or "id" not in df.columns:
        return {}
    result: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        rid = str(row.get("id", "")).strip()
        if rid and rid not in result:
            result[rid] = row.to_dict()
    return result


def merge_contexts(args: argparse.Namespace) -> dict[str, dict[str, Any]]:
    """여러 후보 파일과 검수 sheet의 정보를 id 기준으로 병합한다."""

    selected = row_by_id(read_csv_if_exists(args.selected_csv))
    extension = row_by_id(read_csv_if_exists(args.extension_csv))
    candidates = read_csv_if_exists(args.candidates_csv)
    candidate_groups: dict[str, list[dict[str, Any]]] = {}
    if not candidates.empty and "id" in candidates.columns:
        for _, row in candidates.iterrows():
            candidate_groups.setdefault(str(row.get("id", "")).strip(), []).append(row.to_dict())

    sheet_names = [
        "budget_review",
        "multi_doc_comparison_review",
        "required_fields_review",
        "sub_elig_deadline_review",
        "unanswerable_review",
        "robust_query_review",
        "phase3_extension_candidates",
        "extension_quality_review",
    ]
    sheets = {name: row_by_id(read_workbook_sheet(args.review_xlsx, name)) for name in sheet_names}

    contexts: dict[str, dict[str, Any]] = {}
    all_ids = set(selected) | set(extension) | set(candidate_groups)
    for rid in all_ids:
        ctx: dict[str, Any] = {}
        for source in candidate_groups.get(rid, []):
            for key, value in source.items():
                if key not in ctx or is_blank(ctx[key]):
                    ctx[key] = value
        for source in [selected.get(rid, {}), extension.get(rid, {})]:
            for key, value in source.items():
                if key not in ctx or is_blank(ctx[key]):
                    ctx[key] = value
        for sheet in sheets.values():
            source = sheet.get(rid, {})
            for key, value in source.items():
                if not is_blank(value):
                    ctx[key] = value
        contexts[rid] = ctx
    return contexts


def get_first(row: dict[str, Any], context: dict[str, Any], keys: list[str]) -> Any:
    """row와 context에서 우선순위에 맞는 첫 값을 찾는다."""

    for key in keys:
        value = row.get(key)
        if not is_blank(value):
            return value
    for key in keys:
        value = context.get(key)
        if not is_blank(value):
            return value
    return None


def parse_docs(row: dict[str, Any], context: dict[str, Any]) -> list[str]:
    """source_docs를 정규화한다."""

    value = get_first(
        row,
        context,
        [
            "source_docs",
            "ground_truth_docs_or_source_docs",
            "ground_truth_docs",
            "candidate_source_docs",
            "matched_source_file",
        ],
    )
    docs = as_clean_list(value, max_len=220)
    return [doc for doc in docs if doc]


def parse_evidence(row: dict[str, Any], context: dict[str, Any], source_docs: list[str]) -> list[dict[str, Any]]:
    """근거 정보를 짧은 evidence_refs로 정리한다."""

    value = get_first(row, context, ["evidence_refs", "candidate_fact_chunk_ids", "candidate_budget_fact_chunk_ids"])
    evidence: list[dict[str, Any]] = []
    parsed = as_list(value)
    for item in parsed[:8]:
        if isinstance(item, dict):
            evidence.append(
                {
                    "source_file": clean_text(item.get("source_file") or item.get("source_doc") or "", 220),
                    "chunk_id": clean_text(item.get("chunk_id") or item.get("id") or "", 160),
                    "fact_type": clean_text(item.get("fact_type") or "", 80),
                    "evidence_summary": clean_text(item.get("evidence_summary") or item.get("summary") or "", MAX_EVIDENCE_TEXT),
                }
            )
        else:
            evidence.append(
                {
                    "source_file": source_docs[0] if source_docs else "",
                    "chunk_id": clean_text(item, 160),
                    "fact_type": "",
                    "evidence_summary": "",
                }
            )
    return [ev for ev in evidence if any(ev.values())]


def identity_gold(row: dict[str, Any], context: dict[str, Any]) -> dict[str, list[str]]:
    """발주기관과 사업명을 가능한 범위에서 채운다."""

    agencies = as_clean_list(get_first(row, context, ["reviewer_confirmed_agencies", "candidate_agencies"]), 120)
    projects = as_clean_list(
        get_first(row, context, ["reviewer_confirmed_project_names", "candidate_project_names"]),
        160,
    )
    return {"agencies": agencies, "project_names": projects}


def parse_int(value: Any) -> int | None:
    """문자열에서 KRW 정수 금액을 추출한다."""

    if is_blank(value):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and not math.isnan(value):
        return int(value)
    text = str(value)
    match = re.search(r"-?\d[\d,]*", text)
    if not match:
        return None
    try:
        return int(match.group(0).replace(",", ""))
    except ValueError:
        return None


def build_budget_gold(row: dict[str, Any], context: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    """예산/금액 gold block을 생성한다."""

    raw_items = get_first(row, context, ["reviewer_confirmed_budget_items", "candidate_budget_values"])
    items: list[dict[str, Any]] = []
    for item in as_list(raw_items):
        if isinstance(item, dict):
            amount = parse_int(item.get("amount_krw"))
            source_type = clean_text(item.get("budget_source_type") or "unknown", 80) or "unknown"
            items.append(
                {
                    "label": clean_text(item.get("label") or item.get("project_name") or "", 140),
                    "amount_krw": amount,
                    "budget_source_type": source_type,
                    "source_file": clean_text(item.get("source_file") or "", 220),
                    "chunk_id": clean_text(item.get("chunk_id") or "", 160),
                    "evidence_summary": clean_text(item.get("evidence_summary") or "", MAX_EVIDENCE_TEXT),
                }
            )
            if amount is None:
                warnings.append("budget item 금액을 KRW 정수로 변환하지 못함")
            if source_type == "unknown":
                warnings.append("budget_source_type이 unknown임")
        else:
            amount = parse_int(item)
            if amount is not None:
                items.append(
                    {
                        "label": "",
                        "amount_krw": amount,
                        "budget_source_type": "unknown",
                        "source_file": "",
                        "chunk_id": "",
                        "evidence_summary": "",
                    }
                )
                warnings.append("budget_source_type이 unknown임")
    total = parse_int(get_first(row, context, ["reviewer_confirmed_total_krw", "total_krw"]))
    if total is None and items:
        amounts = [item["amount_krw"] for item in items if item.get("amount_krw") is not None]
        if len(amounts) == len(items):
            total = sum(amounts)
        else:
            warnings.append("일부 금액이 불명확해 total_krw를 확정하지 못함")
    if not items:
        warnings.append("budget 값이 비어 있음")
    excluded = as_clean_list(
        get_first(row, context, ["excluded_budget_candidates", "excluded_budget_candidate_summary"]),
        220,
    )
    return {
        "items": items,
        "total_krw": total,
        "budget_unit": "KRW",
        "tolerance_krw": 0,
        "tolerance_ratio": 0.0,
        "excluded_budget_candidates": excluded,
    }


def make_field(field_name: str, match_type: str, values: list[str], required: bool = True) -> dict[str, Any] | None:
    """required_field_gold의 field 객체를 만든다."""

    values = [clean_text(v, 140) for v in values if clean_text(v, 140)]
    if not values:
        return None
    field: dict[str, Any] = {
        "field_name": field_name,
        "match_type": match_type,
        "required": required,
        "weight": 1.0,
        "evidence_refs": [],
    }
    if match_type in {"keyword_coverage", "checklist_coverage"}:
        field["expected_keywords"] = values
        field["min_match_count"] = min(2, len(values))
    elif len(values) == 1:
        field["expected_value"] = values[0]
    else:
        field["expected_values"] = values
    return field


def build_required_field_gold(row: dict[str, Any], context: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    """필수 정보 gold block을 생성한다."""

    confirmed = get_first(row, context, ["reviewer_confirmed_required_fields"])
    parsed_confirmed = as_list(confirmed)
    if parsed_confirmed and isinstance(parsed_confirmed[0], dict):
        fields = parsed_confirmed
    else:
        fields = []
        for field in [
            make_field("agency", "exact_or_alias", as_clean_list(get_first(row, context, ["candidate_agencies"]), 120)),
            make_field("project_name", "exact_or_alias", as_clean_list(get_first(row, context, ["candidate_project_names"]), 160)),
            make_field("project_duration", "keyword_coverage", as_clean_list(get_first(row, context, ["candidate_project_duration"]), 140)),
            make_field(
                "major_requirements",
                "checklist_coverage",
                as_clean_list(get_first(row, context, ["candidate_major_requirement_keywords"]), 120),
            ),
        ]:
            if field:
                fields.append(field)
    if not fields:
        warnings.append("required_fields인데 fields가 비어 있음")
    return {"fields": fields}


def normalize_deadline(value: Any, warnings: list[str]) -> str | None:
    """날짜 후보를 YYYY-MM-DD 형식으로 정리한다."""

    values = as_clean_list(value, 80)
    if not values:
        return None
    text = values[0]
    match = re.search(r"(20\d{2})[-./년\s]*(\d{1,2}|00)[-./월\s]*(\d{1,2}|00)", text)
    if not match:
        return text
    year, month, day = match.groups()
    if month == "00" or day == "00":
        warnings.append("deadline이 불완전함")
        return None
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def clean_submission_checklist(values: Any) -> list[str]:
    """제출서류 후보에서 긴 안내 문구와 비서류성 잡음을 제거한다."""

    noise_keywords = [
        "손해배상",
        "계약해지",
        "입찰참가자격제한",
        "부정당업자",
        "상호신뢰",
        "반드시 지킬",
        "형사상",
    ]
    result: list[str] = []
    for value in as_list(values):
        text = clean_text(value, 360)
        if not text:
            continue
        parts = re.split(r"[,/|·]+", text)
        for part in parts:
            item = clean_text(part, 80)
            if not item:
                continue
            if len(item) > 45 and any(keyword in item for keyword in noise_keywords):
                continue
            if any(keyword in item for keyword in noise_keywords) and "서약서" not in item:
                continue
            result.append(item)
    return dedupe(result)


def build_submission_gold(row: dict[str, Any], context: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    """제출서류/입찰자격/마감일 gold block을 생성한다."""

    submission = as_clean_list(
        get_first(row, context, ["reviewer_confirmed_submission_documents", "candidate_submission_documents"]),
        140,
    )
    eligibility = as_clean_list(
        get_first(row, context, ["reviewer_confirmed_eligibility_terms", "candidate_eligibility_terms"]),
        180,
    )
    deadline = normalize_deadline(get_first(row, context, ["reviewer_confirmed_deadline", "candidate_deadline"]), warnings)
    deadline_type = clean_text(get_first(row, context, ["deadline_type", "candidate_deadline_type"]) or "", 80)
    if not submission and not eligibility and not deadline:
        warnings.append("submission_documents, eligibility_terms, deadline이 모두 비어 있음")
    if str(get_first(row, context, ["candidate_value_noisy"]) or "").lower() == "true":
        warnings.append("candidate 값에 noisy flag가 있음")
    if str(get_first(row, context, ["question_needs_rewrite"]) or "").lower() == "true":
        warnings.append("question rewrite가 필요한 extension 후보임")
    return {
        "submission_documents": submission,
        "eligibility_terms": eligibility,
        "deadline": deadline,
        "deadline_type": deadline_type,
        "evidence_refs": [],
    }


def build_unanswerable_gold(row: dict[str, Any], context: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    """문서에 없는 질문 gold block을 생성한다."""

    forbidden = as_clean_list(
        get_first(row, context, ["forbidden_claim_types", "forbidden_claim_types_candidate", "unanswerable_reason"]),
        120,
    )
    if not forbidden:
        forbidden = DEFAULT_FORBIDDEN_CLAIM_TYPES
    patterns = as_clean_list(
        get_first(row, context, ["forbidden_hallucination_patterns"]),
        120,
    )
    if not forbidden:
        warnings.append("forbidden_claim_types가 비어 있음")
    return {
        "is_unanswerable": True,
        "allowed_refusal_phrases": ALLOWED_REFUSAL_PHRASES,
        "forbidden_claim_types": forbidden,
        "forbidden_hallucination_patterns": patterns,
    }


def extract_axes_from_question(question: str) -> list[str]:
    """질문에서 과도한 창작 없이 비교축 후보만 추출한다."""

    axes = []
    keywords = {
        "예산": ["예산", "금액", "사업비", "규모"],
        "사업목표": ["목표", "목적"],
        "주요요구사항": ["요구", "기능", "범위", "고도화"],
        "제출/자격": ["제출", "자격", "마감"],
    }
    for label, words in keywords.items():
        if any(word in question for word in words):
            axes.append(label)
    return axes


def build_multi_doc_gold(row: dict[str, Any], context: dict[str, Any], source_docs: list[str], warnings: list[str]) -> dict[str, Any]:
    """다중 문서 비교 gold block을 생성한다."""

    compared_docs = source_docs
    axes = as_clean_list(get_first(row, context, ["candidate_comparison_axes"]), 100)
    if not axes:
        axes = extract_axes_from_question(str(row.get("question", "")))
    output_structure = as_clean_list(get_first(row, context, ["candidate_required_output_structure"]), 80)
    if not output_structure:
        output_structure = DEFAULT_OUTPUT_STRUCTURE
    if len(compared_docs) < 2:
        warnings.append("compared_docs가 2개 미만임")
    if not axes:
        warnings.append("comparison_axes가 비어 있음")
    return {
        "compared_docs": compared_docs,
        "required_doc_coverage": len(compared_docs),
        "required_comparison_axes": axes,
        "required_output_structure": output_structure,
    }


def build_robust_gold(row: dict[str, Any], context: dict[str, Any], source_docs: list[str], warnings: list[str]) -> dict[str, Any]:
    """오타/구어체 견고성 gold block을 생성한다."""

    same_docs = as_clean_list(
        get_first(row, context, ["expected_same_source_docs", "expected_same_source_docs_candidate"]),
        220,
    )
    if not same_docs:
        same_docs = source_docs
    same_fields = as_clean_list(
        get_first(row, context, ["expected_same_key_fields", "expected_same_key_fields_candidate"]),
        120,
    )
    canonical = clean_text(get_first(row, context, ["canonical_question_id", "canonical_question_id_candidate"]) or "", 80)
    related = clean_text(get_first(row, context, ["related_original_id", "related_original_id_candidate"]) or "", 80)
    if not same_docs:
        warnings.append("expected_same_source_docs가 비어 있음")
    if not same_fields:
        warnings.append("expected_same_key_fields가 비어 있음")
    if not canonical and not related:
        warnings.append("canonical_question_id/related_original_id가 모두 비어 있음")
    return {
        "canonical_question_id": canonical or None,
        "related_original_id": related or None,
        "expected_same_source_docs": same_docs,
        "expected_same_key_fields": same_fields,
    }


def build_gold_block(
    row: dict[str, Any],
    context: dict[str, Any],
    source_docs: list[str],
    warnings: list[str],
) -> tuple[str, dict[str, Any]]:
    """task_family에 맞는 gold block을 만든다."""

    task = str(row["primary_task_family"])
    if task == "budget":
        return "budget_gold", build_budget_gold(row, context, warnings)
    if task == "required_fields":
        return "required_field_gold", build_required_field_gold(row, context, warnings)
    if task == "submission_eligibility_deadline":
        return "submission_eligibility_deadline_gold", build_submission_gold(row, context, warnings)
    if task == "unanswerable":
        return "unanswerable_gold", build_unanswerable_gold(row, context, warnings)
    if task == "multi_doc_comparison":
        return "multi_doc_comparison_gold", build_multi_doc_gold(row, context, source_docs, warnings)
    if task == "robust_query_type_e":
        return "robust_query_gold", build_robust_gold(row, context, source_docs, warnings)
    warnings.append(f"알 수 없는 task_family: {task}")
    return "unknown_gold", {}


def status_from_warnings(warnings: list[str], record: dict[str, Any]) -> tuple[str, bool]:
    """warning과 필수값 여부로 생성 상태를 정한다."""

    if not record.get("question") or not record.get("source_docs") or "unknown_gold" in record:
        return "needs_fix", False
    block_name = TASK_BLOCK_NAMES.get(record.get("task_family", ""), "")
    if block_name and block_name not in record:
        return "needs_fix", False
    if warnings:
        return "complete_with_warnings", True
    return "complete", True


def build_record(row: dict[str, Any], context: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    """hybrid_50 한 행을 JSONL record와 사람용 row로 변환한다."""

    warnings: list[str] = []
    source_docs = parse_docs(row, context)
    if not source_docs:
        warnings.append("source_docs가 비어 있음")
    evidence_refs = parse_evidence(row, context, source_docs)
    if not evidence_refs:
        warnings.append("evidence_refs가 비어 있음")
    task = str(row.get("primary_task_family", "")).strip()
    block_name, block = build_gold_block(row, context, source_docs, warnings)
    if block_name == "submission_eligibility_deadline_gold":
        block["evidence_refs"] = evidence_refs
    secondary = as_clean_list(row.get("secondary_task_families"), 100)
    record: dict[str, Any] = {
        "id": str(row.get("id", "")).strip(),
        "source_set": str(row.get("source_set", "")).strip(),
        "question": clean_text(row.get("question"), 400),
        "task_family": task,
        "secondary_task_families": secondary,
        "question_type": str(row.get("type", "")).strip(),
        "difficulty": str(row.get("difficulty", "")).strip(),
        "human_verified": True,
        "review_status": "verified",
        "final_use_decision": "keep",
        "verification_method": BATCH_VERIFICATION_METHOD,
        "reviewer": BATCH_REVIEWER,
        "review_notes": BATCH_REVIEW_NOTES,
        "source_docs": source_docs,
        "identity_gold": identity_gold(row, context),
        "evidence_refs": evidence_refs,
        "notes": clean_text(row.get("selection_reason") or "", 300),
    }
    record[block_name] = block
    status, can_use = status_from_warnings(warnings, record)
    record["gold_generation_status"] = status
    record["gold_generation_warnings"] = dedupe(warnings)
    record["can_use_for_phase3"] = can_use
    readable = build_readable_row(record, block_name)
    validation = {
        "id": record["id"],
        "task_family": task,
        "source_set": record["source_set"],
        "gold_block_type": block_name,
        "gold_generation_status": status,
        "gold_generation_warnings": " | ".join(record["gold_generation_warnings"]),
        "can_use_for_phase3": can_use,
        "has_question": bool(record["question"]),
        "has_source_docs": bool(source_docs),
        "has_evidence_refs": bool(evidence_refs),
        "warning_count": len(record["gold_generation_warnings"]),
    }
    return record, readable, validation


def summarize_values(values: Any, max_len: int = MAX_CELL_TEXT) -> str:
    """사람용 CSV/Excel에 들어갈 요약 문자열을 만든다."""

    if is_blank(values):
        return ""
    if isinstance(values, str):
        return clean_text(values, max_len)
    return clean_text(json.dumps(values, ensure_ascii=False), max_len)


def build_readable_row(record: dict[str, Any], block_name: str) -> dict[str, Any]:
    """사람이 읽기 쉬운 all_gold 행을 만든다."""

    task = record["task_family"]
    block = record.get(block_name, {})
    row = {
        "id": record["id"],
        "question": record["question"],
        "평가유형_한글": TASK_LABELS.get(task, task),
        "task_family": task,
        "source_set": record["source_set"],
        "문항출처_한글": "canonical 500 기반" if record["source_set"] == "canonical_selected_50" else "Phase 3 extension 신규 문항",
        "difficulty": record["difficulty"],
        "source_docs": json_dumps(record["source_docs"]),
        "정답요약_한글": make_answer_summary(task, block),
        "확인해야_하는_핵심값": make_key_values_summary(task, block),
        "gold_block_type": block_name,
        "예산_정답": summarize_values(block if task == "budget" else ""),
        "필수정보_정답": summarize_values(block if task == "required_fields" else ""),
        "제출서류_정답": summarize_values(block.get("submission_documents", "") if task == "submission_eligibility_deadline" else ""),
        "입찰자격_정답": summarize_values(block.get("eligibility_terms", "") if task == "submission_eligibility_deadline" else ""),
        "마감일_정답": summarize_values(block.get("deadline", "") if task == "submission_eligibility_deadline" else ""),
        "답변불가_정답": summarize_values(block if task == "unanswerable" else ""),
        "비교기준_정답": summarize_values(block if task == "multi_doc_comparison" else ""),
        "오타질문_유지조건": summarize_values(block if task == "robust_query_type_e" else ""),
        "evidence_summary": summarize_values(record["evidence_refs"], 500),
        "gold_generation_status": record["gold_generation_status"],
        "gold_generation_warnings": " | ".join(record["gold_generation_warnings"]),
        "can_use_for_phase3": record["can_use_for_phase3"],
        "reviewer": record["reviewer"],
        "verification_method": record["verification_method"],
        "review_notes": record["review_notes"],
    }
    return row


def make_answer_summary(task: str, block: dict[str, Any]) -> str:
    """task_family별 한글 정답 요약을 만든다."""

    if task == "budget":
        total = block.get("total_krw")
        item_count = len(block.get("items", []))
        return f"예산 항목 {item_count}개, total_krw={total}, budget_unit=KRW"
    if task == "multi_doc_comparison":
        return (
            f"{block.get('required_doc_coverage', 0)}개 문서를 비교하고 "
            f"비교축 {', '.join(block.get('required_comparison_axes', [])) or '확인 필요'}를 확인"
        )
    if task == "required_fields":
        names = [field.get("field_name", "") for field in block.get("fields", [])]
        return f"필수 필드 확인: {', '.join(names) or '확인 필요'}"
    if task == "submission_eligibility_deadline":
        parts = []
        if block.get("submission_documents"):
            parts.append("제출서류")
        if block.get("eligibility_terms"):
            parts.append("입찰참가자격")
        if block.get("deadline"):
            parts.append("마감일")
        return f"{', '.join(parts) or '제출/자격/마감 후보 확인 필요'} 확인"
    if task == "unanswerable":
        return "문서에 없는 정보이므로 확인 불가 응답이 정답"
    if task == "robust_query_type_e":
        return "오타/구어체 질문도 같은 정답 문서와 핵심 필드를 유지해야 함"
    return "확인 필요"


def make_key_values_summary(task: str, block: dict[str, Any]) -> str:
    """평가에 직접 쓰이는 핵심값을 짧게 요약한다."""

    if task == "budget":
        items = block.get("items", [])
        source_types = dedupe([item.get("budget_source_type", "") for item in items])
        return f"amount_krw, total_krw={block.get('total_krw')}, budget_source_type={source_types}"
    if task == "multi_doc_comparison":
        return summarize_values(
            {
                "compared_docs": block.get("compared_docs", []),
                "required_comparison_axes": block.get("required_comparison_axes", []),
                "required_output_structure": block.get("required_output_structure", []),
            },
            420,
        )
    if task == "required_fields":
        return summarize_values([field.get("field_name") for field in block.get("fields", [])], 240)
    if task == "submission_eligibility_deadline":
        return summarize_values(
            {
                "submission_documents": block.get("submission_documents", []),
                "eligibility_terms": block.get("eligibility_terms", []),
                "deadline": block.get("deadline"),
            },
            420,
        )
    if task == "unanswerable":
        return summarize_values(
            {
                "allowed_refusal_phrases": block.get("allowed_refusal_phrases", []),
                "forbidden_claim_types": block.get("forbidden_claim_types", []),
            },
            360,
        )
    if task == "robust_query_type_e":
        return summarize_values(
            {
                "expected_same_source_docs": block.get("expected_same_source_docs", []),
                "expected_same_key_fields": block.get("expected_same_key_fields", []),
            },
            360,
        )
    return ""


def build_generation(args: argparse.Namespace) -> GenerationResult:
    """hybrid_50을 최종 gold 산출물 데이터로 변환한다."""

    hybrid = pd.read_csv(args.input_hybrid_csv, encoding="utf-8-sig")
    if len(hybrid) != 50:
        raise ValueError(f"hybrid_50은 50행이어야 합니다. 현재 {len(hybrid)}행입니다.")
    if hybrid["id"].nunique() != 50:
        raise ValueError("hybrid_50 id가 50개 고유값이 아닙니다.")
    contexts = merge_contexts(args)
    records: list[dict[str, Any]] = []
    readable_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []
    for _, hybrid_row in hybrid.iterrows():
        row = hybrid_row.to_dict()
        rid = str(row["id"]).strip()
        context = contexts.get(rid, {})
        record, readable, validation = build_record(row, context)
        records.append(record)
        readable_rows.append(readable)
        validation_rows.append(validation)

    approved = hybrid.copy()
    approved["final_use_decision"] = "keep"
    approved["review_status"] = "verified"
    approved["human_verified"] = True
    approved["verification_method"] = BATCH_VERIFICATION_METHOD
    approved["reviewer"] = BATCH_REVIEWER
    approved["review_notes"] = BATCH_REVIEW_NOTES
    status_by_id = {row["id"]: row for row in validation_rows}
    approved["gold_generation_status"] = approved["id"].map(lambda rid: status_by_id[rid]["gold_generation_status"])
    approved["gold_generation_warnings"] = approved["id"].map(lambda rid: status_by_id[rid]["gold_generation_warnings"])
    approved["can_use_for_phase3"] = approved["id"].map(lambda rid: status_by_id[rid]["can_use_for_phase3"])
    return GenerationResult(records, readable_rows, validation_rows, approved)


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    """JSONL 파일을 저장한다."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    """utf-8-sig CSV를 저장한다."""

    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def add_dataframe_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    """DataFrame 형태의 sheet를 추가한다."""

    ws = wb.create_sheet(name)
    df = pd.DataFrame(rows)
    if df.empty:
        ws.append(["empty"])
    else:
        ws.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws.append([json_dumps(v) if isinstance(v, (list, dict)) else v for v in row])
    style_sheet(ws)


def style_sheet(ws: Any) -> None:
    """Excel sheet 기본 편의 기능을 적용한다."""

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:50]]
        width = min(max([len(value) for value in values] + [10]) + 2, 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_excel(path: Path, readable_rows: list[dict[str, Any]], validation_rows: list[dict[str, Any]]) -> None:
    """사람용 최종 정답지 Excel을 생성한다."""

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    readme_rows = [
        ["항목", "설명"],
        ["파일 목적", "이 Excel은 사람이 읽는 최종 Phase 3 정답지입니다."],
        ["평가 코드 입력", "실제 평가 코드는 rfp_domain_gold_sample.jsonl을 읽습니다."],
        ["사용 용도", "Excel은 검토와 보고서 작성 편의를 위한 파일입니다."],
        ["candidate 값", "candidate_* 값은 자동 후보였고 최종 JSONL에는 gold block으로 변환되었습니다."],
        ["warning", "gold_generation_warnings가 있는 문항은 보고서 작성 시 주의해야 합니다."],
        ["can_use_for_phase3", "false 문항은 공식 평가 전에 확인이 필요합니다."],
        ["보안", "원본 RFP 전문은 포함하지 않습니다."],
        ["evidence", "source_file, chunk_id, 짧은 evidence_summary만 포함합니다."],
    ]
    for row in readme_rows:
        ws.append(row)
    style_sheet(ws)
    add_dataframe_sheet(wb, "all_gold", readable_rows)
    for task in TASK_LABELS:
        task_rows = [row for row in readable_rows if row["task_family"] == task]
        add_dataframe_sheet(wb, task, task_rows)
    add_dataframe_sheet(wb, "validation", validation_rows)
    wb.save(path)


def write_guide(path: Path, record_count: int) -> None:
    """정답지 구조와 영어 필드명을 설명하는 한글 해설서를 작성한다."""

    content = f"""# Phase 3 RFP 도메인 정답지 해설서

## 1. 문서 목적

이 문서는 Phase 3 RFP 특화 평가 정답지인 `rfp_domain_gold_sample.jsonl`의 해설서다. JSONL 파일은 평가 코드가 읽는 파일이고, 이 해설서는 사람이 구조와 영어 필드명을 이해하기 위한 파일이다.

## 2. 최종 정답지 개요

- 정답지 파일: `eval/evaluation/data/rfp_domain_gold_sample.jsonl`
- record 수: {record_count}
- 기반 세트: 사용자가 승인한 `hybrid_50`
- 구성: canonical 문항과 Phase 3 extension 문항이 섞여 있음
- 승인 방식: `batch_user_approval_from_hybrid_50`
- 주의: row-level 수동 검수가 아니라 batch approval이므로 `gold_generation_warnings`를 반드시 확인해야 함
- 사람용 파일: `rfp_domain_gold_sample.xlsx`, `rfp_domain_gold_sample_readable.csv`

## 3. 한눈에 보는 평가 항목

| 평가 항목 | 무엇을 보는지 | gold block | 주요 값 | 주의할 점 |
|---|---|---|---|---|
| 예산/금액 평가 | 금액, 합산, 예산 유형 | `budget_gold` | `amount_krw`, `total_krw`, `budget_source_type` | `threshold_budget`, `payment_terms`를 사업금액으로 오인하지 않음 |
| 여러 문서 비교 평가 | 여러 문서를 모두 다루는지 | `multi_doc_comparison_gold` | `compared_docs`, `required_comparison_axes` | 복잡한 의미 품질은 Phase 4에서 보조 평가 |
| 필수 정보 평가 | 발주기관, 사업명, 사업기간, 주요요구사항 | `required_field_gold` | `fields`, `match_type`, `expected_keywords` | 주요요구사항은 checklist/keyword 중심 |
| 제출서류/입찰자격/마감일 평가 | 제출서류, 자격요건, 날짜 | `submission_eligibility_deadline_gold` | `submission_documents`, `eligibility_terms`, `deadline` | 불완전 날짜는 warning 확인 |
| 문서에 없는 질문 대응 평가 | 확인 불가 응답 여부 | `unanswerable_gold` | `allowed_refusal_phrases`, `forbidden_claim_types` | 문서 밖 단정은 금지 |
| 오타/구어체 견고성 평가 | 같은 정답 문서와 핵심 필드 유지 | `robust_query_gold` | `expected_same_source_docs`, `expected_same_key_fields` | 원 질문 id가 없을 수 있으므로 warning 확인 |

## 4. 영어 필드명 한글 해설표

| 필드명 | 한국어 의미 |
|---|---|
| `id` | 문항 식별자 |
| `source_set` | canonical 문항인지 Phase 3 extension 문항인지 |
| `question` | 평가 질문 |
| `task_family` | Phase 3 평가 유형 |
| `secondary_task_families` | 보조 성격의 평가 유형 |
| `question_type` | 기존 eval type 또는 extension type |
| `difficulty` | 난이도 |
| `human_verified` | 사람 승인 여부 |
| `review_status` | 검수 상태 |
| `final_use_decision` | 최종 사용 결정 |
| `verification_method` | 승인 방식 |
| `reviewer` | 검수자 식별자 |
| `review_notes` | 검수/승인 메모 |
| `source_docs` | 정답 근거 문서명 |
| `identity_gold` | 발주기관, 사업명 등 식별 정보 |
| `evidence_refs` | 짧은 근거 참조 |
| `notes` | 후보 선정 또는 생성 메모 |
| `budget_gold` | 예산/금액 평가용 gold block |
| `required_field_gold` | 필수 정보 평가용 gold block |
| `submission_eligibility_deadline_gold` | 제출서류/입찰자격/마감일 평가용 gold block |
| `unanswerable_gold` | 답변 불가 평가용 gold block |
| `multi_doc_comparison_gold` | 다중 문서 비교 평가용 gold block |
| `robust_query_gold` | 오타/구어체 견고성 평가용 gold block |
| `gold_generation_status` | gold block 생성 상태 |
| `gold_generation_warnings` | 생성 중 확인된 주의사항 |
| `can_use_for_phase3` | Phase 3 공식 평가 사용 가능 여부 |

## 5. gold block별 해설

### identity_gold

발주기관과 사업명을 담는 공통 식별 정보다. 이 값이 있다고 해서 항상 점수화되는 것은 아니며, 실제 점수화 여부는 task별 gold block에서 결정한다.

### budget_gold

사업금액, 합산금액, 예산 종류, 제외해야 할 금액 후보를 담는다. 금액은 가능한 경우 KRW 정수로 저장한다.

### required_field_gold

발주기관, 사업명, 사업기간, 주요요구사항 등 질문에서 요구한 필수 정보를 field 목록으로 담는다.

### submission_eligibility_deadline_gold

제출서류, 입찰참가자격, 제안서 제출기한 또는 입찰 마감일을 담는다. 제출서류와 자격요건은 짧은 checklist 중심이다.

### unanswerable_gold

문서에 없는 질문인지, 어떤 확인 불가 표현을 허용하는지, 어떤 단정을 금지하는지 담는다.

### multi_doc_comparison_gold

비교 대상 문서, 비교 기준, 필요한 출력 구조를 담는다.

### robust_query_gold

오타/구어체 질문이 유지해야 할 정답 문서와 핵심 필드를 담는다.

## 6. JSONL record 예시

### budget 예시

```json
{{"id":"Q-BUDGET","task_family":"budget","budget_gold":{{"items":[{{"label":"A 사업","amount_krw":1000000000,"budget_source_type":"project_budget"}}],"total_krw":1000000000,"budget_unit":"KRW"}}}}
```

### submission_eligibility_deadline 예시

```json
{{"id":"P3-SUB-EXAMPLE","task_family":"submission_eligibility_deadline","submission_eligibility_deadline_gold":{{"submission_documents":["제안서","입찰참가신청서"],"eligibility_terms":["소프트웨어사업자"],"deadline":"2024-09-11"}}}}
```

### unanswerable 예시

```json
{{"id":"Q-UNANSWERABLE","task_family":"unanswerable","unanswerable_gold":{{"is_unanswerable":true,"allowed_refusal_phrases":["확인할 수 없습니다"],"forbidden_claim_types":["낙찰업체 단정"]}}}}
```

## 7. 사람용 Excel/CSV 해설

- `rfp_domain_gold_sample.xlsx`는 사람이 보는 최종 정답지다.
- `rfp_domain_gold_sample_readable.csv`는 빠른 확인용 요약 정답지다.
- Excel의 `all_gold` sheet와 readable CSV는 같은 내용을 담는다.
- task별 sheet는 해당 평가 유형만 필터링해서 보여준다.
- validation sheet는 각 문항의 사용 가능 여부와 warning을 보여준다.

## 8. status와 warning 해설

| 값 | 의미 |
|---|---|
| `complete` | warning 없이 생성됨 |
| `complete_with_warnings` | 사용 가능하지만 보고서에서 주의해야 함 |
| `needs_fix` | 공식 평가 전 확인 필요 |
| `can_use_for_phase3` | Phase 3 평가 투입 가능 여부 |
| `gold_generation_warnings` | 자동 변환 중 부족하거나 불확실했던 값 |

`complete_with_warnings`는 사용할 수 있지만, 결과 해석이나 보고서 작성 시 주의해야 한다. `needs_fix` 또는 `can_use_for_phase3=false`인 문항은 공식 평가에 넣으면 안 된다.

## 9. 보안/데이터 주의사항

- 원본 RFP 전문을 JSONL/XLSX/CSV에 길게 넣지 않는다.
- evidence_refs는 source_file, chunk_id, 짧은 evidence_summary 중심이다.
- API key, 개인정보, secret 값은 포함하지 않는다.
- 이 정답지는 평가용 gold label이지 원본 RFP 대체물이 아니다.

## 10. 다음 단계

1. validation 결과와 warning 문항을 확인한다.
2. `needs_fix` 문항이 있으면 공식 평가 전 보완한다.
3. Phase 3 metric 구현 시 이 JSONL을 입력으로 사용한다.
4. Phase 4 LLM Judge에는 긴 원문이 아니라 짧은 summary만 연결한다.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def write_report(path: Path, result: GenerationResult, cleanup_deleted: list[str], cleanup_candidates: list[Path]) -> None:
    """생성 과정과 검증 결과 보고서를 작성한다."""

    status_counts = Counter(row["gold_generation_status"] for row in result.validation_rows)
    task_counts = Counter(record["task_family"] for record in result.records)
    block_counts = Counter(TASK_BLOCK_NAMES.get(record["task_family"], "unknown_gold") for record in result.records)
    warning_rows = [row for row in result.validation_rows if row["warning_count"]]
    cannot_use = [row for row in result.validation_rows if not row["can_use_for_phase3"]]
    content = [
        "# Phase 3 최종 gold JSONL 생성 리포트",
        "",
        "## 1. 작업 목적",
        "",
        "`hybrid_50` 50문항을 사용자 batch approval에 따라 Phase 3 pilot gold set으로 변환했다.",
        "",
        "## 2. 사용자 batch approval 반영 방식",
        "",
        f"- `human_verified=true`",
        f"- `review_status=verified`",
        f"- `final_use_decision=keep`",
        f"- `verification_method={BATCH_VERIFICATION_METHOD}`",
        f"- `reviewer={BATCH_REVIEWER}`",
        f"- `review_notes={BATCH_REVIEW_NOTES}`",
        "",
        "## 3. record 수와 id 고유성",
        "",
        f"- JSONL record 수: {len(result.records)}",
        f"- id 고유 수: {len({record['id'] for record in result.records})}",
        "",
        "## 4. task_family별 record 수",
        "",
        *[f"- `{task}`: {count}" for task, count in task_counts.items()],
        "",
        "## 5. gold block별 생성 수",
        "",
        *[f"- `{block}`: {count}" for block, count in block_counts.items()],
        "",
        "## 6. gold_generation_status 분포",
        "",
        *[f"- `{status}`: {count}" for status, count in status_counts.items()],
        "",
        "## 7. can_use_for_phase3=false 문항",
        "",
        *(f"- `{row['id']}`: {row['gold_generation_warnings']}" for row in cannot_use),
        "" if cannot_use else "- 없음",
        "",
        "## 8. warning이 있는 문항",
        "",
        *(f"- `{row['id']}`: {row['gold_generation_warnings']}" for row in warning_rows),
        "" if warning_rows else "- 없음",
        "",
        "## 9. 생성 산출물",
        "",
        "- `eval/evaluation/data/rfp_domain_gold_sample.jsonl`",
        "- `eval/evaluation/data/rfp_domain_gold_sample.xlsx`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_readable.csv`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_validation.csv`",
        "- `eval/evaluation/data/rfp_domain_gold_hybrid_50_approved.csv`",
        "- `eval/evaluation/docs/rfp_domain_gold_sample_guide.md`",
        "",
        "## 10. cleanup 정보",
        "",
        f"- 삭제 후보 수: {len(cleanup_candidates)}",
        f"- 삭제한 파일 수: {len(cleanup_deleted)}",
        *(f"- 삭제: `{item}`" for item in cleanup_deleted),
        "",
        "## 11. 주의사항",
        "",
        "없는 정답값은 새로 만들지 않았고, 부족하거나 불확실한 값은 warning으로 남겼다. 원본 RFP 전문, 긴 table, source_store 전체, secret 값은 산출물에 넣지 않았다.",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(content), encoding="utf-8")


def block_name_for_record(record: dict[str, Any]) -> str:
    """record 안에 들어 있는 gold block 이름을 찾는다."""

    for block_name in [
        "budget_gold",
        "multi_doc_comparison_gold",
        "required_field_gold",
        "submission_eligibility_deadline_gold",
        "unanswerable_gold",
        "robust_query_gold",
    ]:
        if block_name in record:
            return block_name
    return "unknown_gold"


def make_limited_evidence_refs(source_docs: list[str]) -> list[dict[str, Any]]:
    """chunk_id를 확정할 수 없을 때 source_docs 기준 제한 근거를 만든다."""

    return [
        {
            "source_file": source_doc,
            "chunk_id": None,
            "fact_type": "",
            "evidence_summary": "source_docs 기준 근거 문서 확인. 세부 chunk 미확정",
        }
        for source_doc in source_docs
    ]


def evidence_keywords_for_record(record: dict[str, Any]) -> list[str]:
    """task_family와 gold block에서 evidence 검색용 키워드를 만든다."""

    task = record.get("task_family")
    keywords: list[str] = []
    if task == "budget":
        keywords.extend(["project_budget", "사업금액", "사업비", "예산", "금액"])
        for item in record.get("budget_gold", {}).get("items", []):
            amount = item.get("amount_krw")
            if amount:
                keywords.append(str(amount))
                keywords.append(f"{amount:,}")
            if item.get("label"):
                keywords.append(str(item["label"]))
    elif task == "multi_doc_comparison":
        keywords.extend(["document_summary", "사업개요", "사업목표", "주요요구사항", "기대효과"])
        keywords.extend(record.get("multi_doc_comparison_gold", {}).get("required_comparison_axes", []))
    elif task == "required_fields":
        keywords.extend(["document_summary", "사업기간", "주요요구사항", "사업개요", "사업목표"])
        for field in record.get("required_field_gold", {}).get("fields", []):
            keywords.append(str(field.get("field_name", "")))
            keywords.extend(field.get("expected_keywords", []))
            if field.get("expected_value"):
                keywords.append(str(field["expected_value"]))
            keywords.extend(field.get("expected_values", []))
    elif task == "submission_eligibility_deadline":
        keywords.extend(["submission_documents", "eligibility", "bid_deadline", "제출서류", "입찰참가자격", "마감"])
    else:
        keywords.extend(["document_summary", "사업개요", "핵심"])
    return [clean_text(keyword, 80) for keyword in keywords if clean_text(keyword, 80)]


def chunk_score(chunk: dict[str, Any], keywords: list[str]) -> int:
    """chunk가 warning record의 근거로 얼마나 적합한지 점수화한다."""

    content = chunk.get("content") or ""
    metadata = chunk.get("metadata") or {}
    fact_type = metadata.get("fact_type") or ""
    score = 0
    if chunk.get("chunk_type") == "fact_candidates":
        score += 5
    if fact_type:
        score += 3
    for keyword in keywords:
        if keyword and keyword in content:
            score += 2
        if keyword and keyword == fact_type:
            score += 5
    return score


def make_chunk_evidence(chunk: dict[str, Any]) -> dict[str, Any]:
    """chunk 객체를 짧은 evidence_ref로 변환한다."""

    metadata = chunk.get("metadata") or {}
    content = clean_text(chunk.get("content") or "", MAX_EVIDENCE_TEXT)
    return {
        "source_file": clean_text(chunk.get("source_file") or metadata.get("source_file") or "", 220),
        "chunk_id": clean_text(chunk.get("chunk_id") or "", 160),
        "fact_type": clean_text(metadata.get("fact_type") or chunk.get("chunk_type") or "", 80),
        "evidence_summary": content,
    }


def find_evidence_for_records(records: list[dict[str, Any]], chunks_path: Path) -> dict[str, list[dict[str, Any]]]:
    """chunks_v2에서 warning 문항의 evidence_refs 후보를 찾는다."""

    targets = {
        record["id"]: record
        for record in records
        if record.get("gold_generation_warnings") and not record.get("evidence_refs")
    }
    if not targets or not chunks_path.exists():
        return {}
    source_to_ids: dict[str, list[str]] = {}
    for record in targets.values():
        for source_doc in record.get("source_docs", []):
            source_to_ids.setdefault(source_doc, []).append(record["id"])
    best: dict[tuple[str, str], tuple[int, dict[str, Any]]] = {}
    keywords_by_id = {rid: evidence_keywords_for_record(record) for rid, record in targets.items()}
    with chunks_path.open("r", encoding="utf-8") as f:
        for line in f:
            chunk = json.loads(line)
            source_file = chunk.get("source_file") or (chunk.get("metadata") or {}).get("source_file")
            if source_file not in source_to_ids:
                continue
            for rid in source_to_ids[source_file]:
                score = chunk_score(chunk, keywords_by_id[rid])
                key = (rid, source_file)
                if score > best.get(key, (-1, {}))[0]:
                    best[key] = (score, chunk)
    result: dict[str, list[dict[str, Any]]] = {}
    for (rid, _source), (score, chunk) in best.items():
        if score <= 0:
            continue
        result.setdefault(rid, []).append(make_chunk_evidence(chunk))
    return result


def normalize_doc_name(value: str) -> str:
    """문서명 비교를 위해 공백과 특수문자를 제거한다."""

    return re.sub(r"[^0-9A-Za-z가-힣]+", "", value or "").lower()


def load_eval_rows(eval_dir: Path) -> list[dict[str, Any]]:
    """canonical eval CSV 25개를 읽는다."""

    rows: list[dict[str, Any]] = []
    for idx in range(1, 26):
        path = eval_dir / f"eval_batch_{idx:02d}.csv"
        if not path.exists():
            continue
        df = pd.read_csv(path, encoding="utf-8-sig")
        rows.extend(df.to_dict("records"))
    return rows


def parse_eval_docs(value: Any) -> list[str]:
    """eval CSV의 ground_truth_docs를 리스트로 파싱한다."""

    return as_clean_list(value, 240)


def infer_key_fields_from_question(question: str) -> list[str]:
    """robust 질문에서 유지해야 할 핵심 필드 후보를 추출한다."""

    fields = []
    if any(word in question for word in ["예산", "에산", "예싼", "금액"]):
        fields.append("budget")
    if any(word in question for word in ["기대효과", "기대효괍", "효과"]):
        fields.append("expected_effects")
    if any(word in question for word in ["사업", "플젝", "시스템"]):
        fields.append("project_identity")
    if not fields:
        fields = ["정답 문서", "핵심 필드"]
    return dedupe(fields)


def find_related_original_id(record: dict[str, Any], eval_rows: list[dict[str, Any]]) -> str | None:
    """robust 문항과 같은 문서/핵심 의도를 가진 원 질문 후보를 찾는다."""

    source_norms = {normalize_doc_name(doc) for doc in record.get("source_docs", [])}
    if not source_norms:
        return None
    question = str(record.get("question", ""))
    key_fields = infer_key_fields_from_question(question)
    best_score = 0.0
    best_id: str | None = None
    for row in eval_rows:
        rid = str(row.get("id", ""))
        if rid == record.get("id"):
            continue
        row_docs = {normalize_doc_name(doc) for doc in parse_eval_docs(row.get("ground_truth_docs"))}
        if not source_norms.intersection(row_docs):
            continue
        # robust 원 질문은 같은 정답 문서 집합일 때만 확정한다.
        # extra 문서가 포함된 다중 문서 질문은 참고 후보일 수 있지만 원 질문으로 확정하지 않는다.
        if source_norms != row_docs:
            continue
        row_question = str(row.get("question", ""))
        score = difflib.SequenceMatcher(None, normalize_doc_name(question), normalize_doc_name(row_question)).ratio()
        if "budget" in key_fields and any(word in row_question for word in ["예산", "금액", "사업비"]):
            score += 0.35
        if "expected_effects" in key_fields and any(word in row_question for word in ["기대효과", "성과", "효과", "목표"]):
            score += 0.35
        if row.get("type") == "E":
            score -= 0.2
        if score > best_score:
            best_score = score
            best_id = rid
    if best_score >= 0.45:
        return best_id
    return None


def remove_warning(record: dict[str, Any], warning_text: str) -> None:
    """특정 warning을 제거한다."""

    record["gold_generation_warnings"] = [
        warning for warning in record.get("gold_generation_warnings", []) if warning != warning_text
    ]


def refresh_record_status(record: dict[str, Any]) -> None:
    """warning 상태에 따라 status/can_use 값을 갱신한다."""

    warnings = record.get("gold_generation_warnings", [])
    resolution = record.get("warning_resolution_status")
    if resolution == "unresolved_needs_fix":
        record["gold_generation_status"] = "needs_fix"
        record["can_use_for_phase3"] = False
    elif warnings:
        record["gold_generation_status"] = "complete_with_warnings"
        record["can_use_for_phase3"] = True
    else:
        record["gold_generation_status"] = "complete"
        record["can_use_for_phase3"] = True


def resolve_warning_records(records: list[dict[str, Any]], chunks_path: Path, eval_dir: Path) -> dict[str, Any]:
    """14개 warning 문항을 해결 또는 accepted_warning으로 분류한다."""

    old_warning_ids = [
        record["id"]
        for record in records
        if record.get("gold_generation_warnings") or record.get("warning_resolution_status")
    ]
    evidence_hits = find_evidence_for_records(records, chunks_path)
    eval_rows = load_eval_rows(eval_dir)
    resolved: list[str] = []
    accepted: list[str] = []
    unresolved: list[str] = []
    notes_by_id: dict[str, str] = {}

    for record in records:
        rid = record["id"]
        old_warnings = list(record.get("gold_generation_warnings", []))
        if not old_warnings and rid in {"Q019", "Q020", "Q039"} and record.get("warning_resolution_status"):
            old_warnings = ["canonical_question_id/related_original_id가 모두 비어 있음"]
        if not old_warnings:
            if record.get("warning_resolution_status"):
                continue
            record["warning_resolution_status"] = ""
            record["warning_resolution_notes"] = ""
            continue
        notes: list[str] = []
        status = "accepted_warning"

        if "evidence_refs가 비어 있음" in old_warnings:
            found = evidence_hits.get(rid, [])
            if found:
                record["evidence_refs"] = found
                block_name = block_name_for_record(record)
                if block_name == "submission_eligibility_deadline_gold":
                    record[block_name]["evidence_refs"] = found
                remove_warning(record, "evidence_refs가 비어 있음")
                status = "resolved"
                notes.append("new_data/chunks_v2_690.jsonl에서 source_file 기준 fact/text chunk 근거를 보완함")
            else:
                record["evidence_refs"] = make_limited_evidence_refs(record.get("source_docs", []))
                record["gold_generation_warnings"] = [
                    warning for warning in old_warnings if warning != "evidence_refs가 비어 있음"
                ]
                record["gold_generation_warnings"].append("세부 chunk_id 미확정 limited evidence_ref 사용")
                status = "accepted_warning"
                notes.append("정확한 chunk_id는 찾지 못했으나 source_docs가 명확해 제한 근거로 사용 가능함")

        if rid == "P3-SUB-001":
            block = record.get("submission_eligibility_deadline_gold", {})
            block["submission_documents"] = clean_submission_checklist(block.get("submission_documents", []))
            block["eligibility_terms"] = []
            remove_warning(record, "candidate 값에 noisy flag가 있음")
            status = "resolved"
            notes.append("질문 의도에 맞춰 제출서류 checklist만 남기고 noisy eligibility 문구를 제거함")

        if rid == "P3-SUB-004":
            original = record.get("question", "")
            record["original_question"] = original
            record["question"] = "경기도 포천시의 2024년 포천시 스마트도시 솔루션 확산사업 입찰참가자격은 무엇입니까?"
            remove_warning(record, "question rewrite가 필요한 extension 후보임")
            status = "resolved"
            notes.append("입찰참가자격 중심으로 질문 문장을 자연스럽게 수정함")

        if rid == "P3-SUB-010":
            original = record.get("question", "")
            record["original_question"] = original
            record["question"] = "대전대학교 2024학년도 다층적 융합 학습경험 플랫폼(MILE) 사업의 제출서류, 입찰참가자격, 입찰마감일은 무엇입니까?"
            remove_warning(record, "question rewrite가 필요한 extension 후보임")
            status = "resolved"
            notes.append("gold block에 실제 포함된 제출서류/입찰참가자격/마감일 범위로 질문을 좁힘")

        if rid in {"Q019", "Q020", "Q039"}:
            block = record.get("robust_query_gold", {})
            related = find_related_original_id(record, eval_rows)
            if related:
                block["related_original_id"] = related
                block["canonical_question_id"] = related
                remove_warning(record, "canonical_question_id/related_original_id가 모두 비어 있음")
                status = "resolved"
                notes.append(f"canonical eval에서 같은 source_docs와 유사 의도를 가진 원 질문 후보 {related}를 연결함")
            else:
                block["related_original_id"] = None
                block["canonical_question_id"] = None
                block["expected_same_source_docs"] = block.get("expected_same_source_docs") or record.get("source_docs", [])
                block["expected_same_key_fields"] = block.get("expected_same_key_fields") or infer_key_fields_from_question(record.get("question", ""))
                record["gold_generation_warnings"] = [
                    warning for warning in old_warnings if warning != "canonical_question_id/related_original_id가 모두 비어 있음"
                ]
                record["gold_generation_warnings"].append(
                    "related_original_id 미확정이나 same source/key field 기준으로 평가 가능"
                )
                status = "accepted_warning"
                notes.append("신뢰도 높은 원 질문 id를 확정하지 못했으나 same source/key field 기준 robust 평가가 가능함")

        record["warning_resolution_status"] = status
        record["warning_resolution_notes"] = " ".join(notes) if notes else "warning 검토 완료"
        refresh_record_status(record)
        if status == "resolved":
            resolved.append(rid)
        elif status == "accepted_warning":
            accepted.append(rid)
        else:
            unresolved.append(rid)
        notes_by_id[rid] = record["warning_resolution_notes"]

    resolved = [
        record["id"] for record in records if record.get("warning_resolution_status") == "resolved"
    ]
    accepted = [
        record["id"] for record in records if record.get("warning_resolution_status") == "accepted_warning"
    ]
    unresolved = [
        record["id"] for record in records if record.get("warning_resolution_status") == "unresolved_needs_fix"
    ]
    notes_by_id = {
        record["id"]: record.get("warning_resolution_notes", "")
        for record in records
        if record.get("warning_resolution_status")
    }
    return {
        "old_warning_ids": old_warning_ids,
        "resolved": resolved,
        "accepted_warning": accepted,
        "unresolved_needs_fix": unresolved,
        "notes_by_id": notes_by_id,
    }


def rows_from_records(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """기존 JSONL record에서 readable/validation rows를 다시 만든다."""

    readable_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []
    for record in records:
        block_name = block_name_for_record(record)
        readable = build_readable_row(record, block_name)
        readable["warning_resolution_status"] = record.get("warning_resolution_status", "")
        readable["warning_resolution_notes"] = record.get("warning_resolution_notes", "")
        readable_rows.append(readable)
        validation_rows.append(
            {
                "id": record["id"],
                "task_family": record["task_family"],
                "source_set": record["source_set"],
                "gold_block_type": block_name,
                "gold_generation_status": record["gold_generation_status"],
                "gold_generation_warnings": " | ".join(record.get("gold_generation_warnings", [])),
                "can_use_for_phase3": record["can_use_for_phase3"],
                "has_question": bool(record.get("question")),
                "has_source_docs": bool(record.get("source_docs")),
                "has_evidence_refs": bool(record.get("evidence_refs")),
                "warning_count": len(record.get("gold_generation_warnings", [])),
                "warning_resolution_status": record.get("warning_resolution_status", ""),
                "warning_resolution_notes": record.get("warning_resolution_notes", ""),
            }
        )
    return readable_rows, validation_rows


def update_approved_csv_for_resolution(path: Path, records: list[dict[str, Any]]) -> None:
    """approved CSV에 warning resolution 결과를 반영한다."""

    if path.exists():
        df = pd.read_csv(path, encoding="utf-8-sig")
    else:
        df = pd.DataFrame([{"id": record["id"]} for record in records])
    by_id = {record["id"]: record for record in records}
    for col in [
        "question",
        "gold_generation_status",
        "gold_generation_warnings",
        "can_use_for_phase3",
        "warning_resolution_status",
        "warning_resolution_notes",
    ]:
        df[col] = df["id"].map(
            lambda rid: (
                " | ".join(by_id[rid].get(col, []))
                if col == "gold_generation_warnings"
                else by_id[rid].get(col, "")
            )
        )
    df.to_csv(path, index=False, encoding="utf-8-sig")


def append_resolution_to_generation_report(path: Path, summary: dict[str, Any], status_counts: Counter[str]) -> None:
    """기존 생성 리포트에 warning resolution 섹션을 추가한다."""

    section = [
        "",
        "## Warning resolution pass",
        "",
        "Phase 3 Gold Warning Resolution 작업을 수행했다.",
        "",
        f"- 기존 warning 문항 수: {len(summary['old_warning_ids'])}",
        f"- resolved: {len(summary['resolved'])}",
        f"- accepted_warning: {len(summary['accepted_warning'])}",
        f"- unresolved_needs_fix: {len(summary['unresolved_needs_fix'])}",
        f"- 최종 status 분포: {dict(status_counts)}",
        "- can_use_for_phase3=false 문항: 없음" if not summary["unresolved_needs_fix"] else f"- can_use_for_phase3=false 문항: {summary['unresolved_needs_fix']}",
    ]
    text = path.read_text(encoding="utf-8") if path.exists() else "# Phase 3 최종 gold JSONL 생성 리포트\n"
    marker = "## Warning resolution pass"
    if marker in text:
        text = text.split(marker)[0].rstrip()
    path.write_text(text.rstrip() + "\n" + "\n".join(section) + "\n", encoding="utf-8")


def update_guide_for_resolution(path: Path) -> None:
    """해설서에 warning resolution 필드 설명을 추가한다."""

    text = path.read_text(encoding="utf-8")
    if "warning_resolution_status" in text:
        return
    insert = """

## Warning resolution 필드 해설

| 필드명 | 한국어 의미 |
|---|---|
| `warning_resolution_status` | warning 처리 상태. `resolved`, `accepted_warning`, `unresolved_needs_fix` 중 하나 |
| `warning_resolution_notes` | warning을 어떻게 처리했는지에 대한 짧은 설명 |

`resolved`는 근거 보완, 질문 rewrite, noisy 값 정리 등으로 warning이 실제로 해결된 상태다. `accepted_warning`은 warning이 남아 있지만 Phase 3 정답지 사용에는 문제가 없다고 판단한 상태다. 예를 들어 robust 원 질문 id가 없어도 `expected_same_source_docs`와 `expected_same_key_fields`가 충분하면 accepted_warning으로 둘 수 있다.

`unresolved_needs_fix`는 공식 평가 전에 반드시 사람이 고쳐야 하는 상태이며, 이 경우 `can_use_for_phase3=false`로 처리해야 한다.
"""
    path.write_text(text.rstrip() + insert + "\n", encoding="utf-8")


def write_warning_resolution_report(
    path: Path,
    summary: dict[str, Any],
    records: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
) -> None:
    """warning resolution 결과 보고서를 작성한다."""

    status_counts = Counter(row["gold_generation_status"] for row in validation_rows)
    resolution_counts = Counter(
        record.get("warning_resolution_status")
        for record in records
        if record.get("warning_resolution_status")
    )
    false_ids = [row["id"] for row in validation_rows if not row["can_use_for_phase3"]]
    warning_rows = [row for row in validation_rows if row["warning_count"] > 0]
    lines = [
        "# Phase 3 Gold Warning Resolution Report",
        "",
        "## 1. 작업 목적",
        "",
        "최종 Phase 3 gold set의 warning 14개를 해결 가능한 범위에서 보완하고, 남는 warning은 accepted_warning으로 명확히 분류했다.",
        "",
        "## 2. 입력 파일 목록",
        "",
        "- `eval/evaluation/data/rfp_domain_gold_sample.jsonl`",
        "- `eval/evaluation/data/rfp_domain_gold_sample.xlsx`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_readable.csv`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_validation.csv`",
        "- `new_data/chunks_v2_690.jsonl`",
        "- `data/eval/eval_batch_01.csv` ~ `eval_batch_25.csv`",
        "",
        "## 3. 기존 warning 14개 요약",
        "",
        *[f"- `{rid}`" for rid in summary["old_warning_ids"]],
        "",
        "## 4. warning 유형별 처리 전략",
        "",
        "- evidence_refs 없음: chunks_v2에서 source_file 기준 근거 chunk를 찾고, 실패 시 limited evidence_ref로 accepted_warning 처리",
        "- noisy candidate: 제출서류 checklist만 남기고 긴 안내/계약 일반 문구 제거",
        "- question rewrite: 질문 의미와 gold block 범위를 유지하면서 자연스럽게 수정",
        "- robust 원 질문 id 없음: canonical eval에서 같은 source_docs와 유사 의도 질문을 찾고, 불명확하면 accepted_warning 처리",
        "",
        "## 5. resolved 문항 목록",
        "",
        *[f"- `{rid}`: {summary['notes_by_id'].get(rid, '')}" for rid in summary["resolved"]],
        "",
        "## 6. accepted_warning 문항 목록",
        "",
        *[f"- `{rid}`: {summary['notes_by_id'].get(rid, '')}" for rid in summary["accepted_warning"]],
        "" if summary["accepted_warning"] else "- 없음",
        "",
        "## 7. unresolved_needs_fix 문항 목록",
        "",
        *[f"- `{rid}`" for rid in summary["unresolved_needs_fix"]],
        "" if summary["unresolved_needs_fix"] else "- 없음",
        "",
        "## 8. evidence_refs 보완 결과",
        "",
        "evidence_refs가 비어 있던 문항은 chunks_v2 검색 또는 limited evidence_ref 방식으로 보완했다.",
        "",
        "## 9. P3-SUB-001 noisy candidate 처리 결과",
        "",
        "제출서류 checklist만 남기고, 질문 의도와 직접 관련이 낮은 긴 eligibility/계약 일반 문구는 제거했다.",
        "",
        "## 10. P3-SUB-004/P3-SUB-010 question rewrite 결과",
        "",
        "- `P3-SUB-004`: 입찰참가자격 중심 질문으로 수정",
        "- `P3-SUB-010`: 제출서류, 입찰참가자격, 입찰마감일 범위로 수정",
        "",
        "## 11. Q019/Q020/Q039 robust 연결 id 처리 결과",
        "",
        "canonical eval에서 같은 source_docs와 유사 의도 질문 후보를 탐색했다. 신뢰 가능한 경우 related_original_id를 채웠고, 불명확한 경우에는 same source/key field 기준으로 accepted_warning 처리한다.",
        "",
        "## 12. 최종 gold_generation_status 분포",
        "",
        *[f"- `{key}`: {value}" for key, value in status_counts.items()],
        "",
        "## 13. can_use_for_phase3=false 문항 목록",
        "",
        *(f"- `{rid}`" for rid in false_ids),
        "" if false_ids else "- 없음",
        "",
        "## 14. 정답지 사용 가능성 판단",
        "",
        "unresolved_needs_fix가 없고 can_use_for_phase3=false 문항이 없으므로 Phase 3 metric 구현 단계로 넘어갈 수 있다. accepted_warning 문항은 보고서에서 주의사항으로 표시하면 된다.",
        "",
        "## 15. 아직 남은 리스크",
        "",
        *(f"- `{row['id']}`: {row['gold_generation_warnings']}" for row in warning_rows),
        "" if warning_rows else "- 남은 warning 없음",
        "",
        "## 16. 다음 단계 제안",
        "",
        "Phase 3 metric 구현 시 `warning_resolution_status`를 결과 리포트에 함께 표시하고, accepted_warning 문항은 평가 제외가 아니라 해석 주의 문항으로 다룬다.",
        "",
        "## resolution status 분포",
        "",
        *[f"- `{key}`: {value}" for key, value in resolution_counts.items()],
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")

def update_review_workbook(path: Path, approved_df: pd.DataFrame) -> None:
    """review workbook의 hybrid_50_recommendation sheet에 승인 상태를 반영한다."""

    if not path.exists():
        return
    wb = load_workbook(path)
    if "hybrid_50_recommendation" not in wb.sheetnames:
        return
    ws = wb["hybrid_50_recommendation"]
    headers = [cell.value for cell in ws[1]]
    required_cols = [
        "final_use_decision",
        "review_status",
        "human_verified",
        "verification_method",
        "reviewer",
        "review_notes",
        "gold_generation_status",
        "gold_generation_warnings",
        "can_use_for_phase3",
    ]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
    col_index = {header: idx + 1 for idx, header in enumerate(headers)}
    by_id = {str(row["id"]): row for _, row in approved_df.iterrows()}
    for row_idx in range(2, ws.max_row + 1):
        rid = str(ws.cell(row=row_idx, column=col_index["id"]).value)
        data = by_id.get(rid)
        if data is None:
            continue
        for col in required_cols:
            ws.cell(row=row_idx, column=col_index[col], value=data[col])
    wb.save(path)


def validate_outputs(args: argparse.Namespace, result: GenerationResult) -> dict[str, Any]:
    """생성된 산출물을 다시 읽어 필수 조건을 검증한다."""

    jsonl_records = []
    with args.output_jsonl.open("r", encoding="utf-8") as f:
        for line in f:
            jsonl_records.append(json.loads(line))
    readable = pd.read_csv(args.output_readable_csv, encoding="utf-8-sig")
    wb = load_workbook(args.output_xlsx, read_only=True, data_only=True)
    all_gold_rows = wb["all_gold"].max_row - 1
    ids = [record["id"] for record in jsonl_records]
    text_blob = args.output_jsonl.read_text(encoding="utf-8")
    secret_hits = [pattern.pattern for pattern in SECRET_PATTERNS if pattern.search(text_blob)]
    long_evidence = []
    for record in jsonl_records:
        for evidence in record.get("evidence_refs", []):
            if len(str(evidence.get("evidence_summary", ""))) > MAX_EVIDENCE_TEXT + 20:
                long_evidence.append(record["id"])
    return {
        "jsonl_exists": args.output_jsonl.exists(),
        "xlsx_exists": args.output_xlsx.exists(),
        "readable_csv_exists": args.output_readable_csv.exists(),
        "guide_exists": args.output_guide.exists(),
        "jsonl_count": len(jsonl_records),
        "unique_id_count": len(set(ids)),
        "all_gold_rows": all_gold_rows,
        "readable_rows": len(readable),
        "all_have_question": all(bool(record.get("question")) for record in jsonl_records),
        "all_have_task_family": all(bool(record.get("task_family")) for record in jsonl_records),
        "all_human_verified": all(record.get("human_verified") is True for record in jsonl_records),
        "all_review_verified": all(record.get("review_status") == "verified" for record in jsonl_records),
        "all_keep": all(record.get("final_use_decision") == "keep" for record in jsonl_records),
        "all_batch_method": all(record.get("verification_method") == BATCH_VERIFICATION_METHOD for record in jsonl_records),
        "all_have_source_docs": all(bool(record.get("source_docs")) for record in jsonl_records),
        "readable_has_question": "question" in readable.columns and readable["question"].notna().all(),
        "readable_has_korean_columns": all(
            col in readable.columns for col in ["평가유형_한글", "정답요약_한글", "확인해야_하는_핵심값"]
        ),
        "secret_hits": secret_hits,
        "long_evidence_ids": sorted(set(long_evidence)),
        "sheets": wb.sheetnames,
    }


def cleanup_candidates(args: argparse.Namespace) -> list[Path]:
    """삭제 후보 파일 목록을 만든다."""

    data_dir = args.output_jsonl.parent
    docs_dir = args.output_report.parent
    return [
        data_dir / "rfp_domain_gold_candidates.csv",
        data_dir / "rfp_domain_gold_selected_50.csv",
        data_dir / "rfp_domain_gold_phase3_extension_candidates.csv",
        data_dir / "rfp_domain_gold_hybrid_50_recommendation.csv",
        data_dir / "rfp_domain_gold_review.xlsx",
        docs_dir / "phase3_candidate_generation_report.md",
        docs_dir / "phase3_selected_50_report.md",
        docs_dir / "phase3_extension_candidate_report.md",
        docs_dir / "phase3_extension_quality_and_hybrid_report.md",
        docs_dir / "phase3_hybrid_50_review_checklist.md",
        docs_dir / "phase3_hybrid_50_review_preparation_report.md",
        docs_dir / "phase3_human_review_ux_update_report.md",
        args.scripts_dir / "build_phase3_gold_review_assets.py",
        args.scripts_dir / "select_phase3_gold_50.py",
        args.scripts_dir / "build_phase3_extension_candidates.py",
        args.scripts_dir / "review_phase3_extension_quality_and_hybrid.py",
    ]


def write_cleanup_manifest(
    path: Path,
    args: argparse.Namespace,
    validation: dict[str, Any],
    cleanup_deleted: list[str],
    candidates: list[Path],
    cleanup_enabled: bool,
) -> None:
    """보존/삭제 파일 목록을 manifest에 기록한다."""

    preserve = [
        args.output_jsonl,
        args.output_xlsx,
        args.output_readable_csv,
        args.output_validation_csv,
        args.output_approved_csv,
        args.output_guide,
        args.output_report,
        args.output_cleanup_manifest,
        args.scripts_dir / "build_phase3_gold_sample_jsonl.py",
        Path("eval/evaluation/README.md"),
        Path("eval/evaluation/AGENTS.md"),
        Path("eval/evaluation/src"),
        Path("eval/evaluation/scripts/run_evaluation.py"),
        Path("eval/evaluation/tests"),
        Path("eval/evaluation/requirements.txt"),
        Path("eval/evaluation/docs/evaluation_plan.md"),
        Path("eval/evaluation/docs/evaluation_refactor_plan.md"),
        Path("eval/evaluation/docs/phase3_phase4_eval_plan.md"),
    ]
    not_deleted = [str(path) for path in candidates if path.exists()]
    content = [
        "# Phase 3 cleanup manifest",
        "",
        "## 1. 정리 목적",
        "",
        "Phase 3 최종 gold 산출물이 생성된 뒤 중간 후보 파일을 정리해 평가 폴더를 단순화한다.",
        "",
        "## 2. 보존 파일 목록",
        "",
        *[f"- `{item}`" for item in preserve],
        "",
        "## 3. 삭제한 파일 목록",
        "",
        *(f"- `{item}`" for item in cleanup_deleted),
        "" if cleanup_deleted else "- 없음",
        "",
        "## 4. 삭제하지 않은 파일 목록과 이유",
        "",
        *[f"- `{item}`: cleanup 미실행 또는 파일 미존재" for item in not_deleted],
        "" if not_deleted else "- 삭제 후보 파일이 남아 있지 않음",
        "",
        "## 5. 삭제 전 검증 결과",
        "",
        f"- cleanup 실행 여부: {cleanup_enabled}",
        f"- JSONL record 수: {validation.get('jsonl_count')}",
        f"- id 고유 수: {validation.get('unique_id_count')}",
        f"- Excel all_gold 행 수: {validation.get('all_gold_rows')}",
        f"- readable CSV 행 수: {validation.get('readable_rows')}",
        f"- secret 의심 패턴: {validation.get('secret_hits')}",
        f"- 긴 evidence 의심 id: {validation.get('long_evidence_ids')}",
        "",
        "## 6. 최종 남은 Phase 3 핵심 파일",
        "",
        "- `eval/evaluation/data/rfp_domain_gold_sample.jsonl`",
        "- `eval/evaluation/data/rfp_domain_gold_sample.xlsx`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_readable.csv`",
        "- `eval/evaluation/data/rfp_domain_gold_sample_validation.csv`",
        "- `eval/evaluation/data/rfp_domain_gold_hybrid_50_approved.csv`",
        "- `eval/evaluation/docs/rfp_domain_gold_sample_guide.md`",
        "- `eval/evaluation/docs/phase3_gold_jsonl_generation_report.md`",
        "- `eval/evaluation/scripts/build_phase3_gold_sample_jsonl.py`",
        "",
        "## 7. 주의사항",
        "",
        "평가 실행에 필요한 `run_evaluation.py`, `src`, `tests`, `README.md`, `AGENTS.md`는 삭제하지 않는다.",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(content), encoding="utf-8")


def run_cleanup(args: argparse.Namespace, validation: dict[str, Any], result: GenerationResult) -> list[str]:
    """cleanup 옵션이 있으면 삭제 후보 파일을 정리한다."""

    candidates = cleanup_candidates(args)
    if not args.cleanup:
        return []
    if any(row["gold_generation_status"] == "needs_fix" for row in result.validation_rows):
        print("needs_fix 문항이 있어 cleanup을 실행하지 않습니다.", file=sys.stderr)
        return []
    required_ok = (
        validation["jsonl_count"] == 50
        and validation["unique_id_count"] == 50
        and validation["all_gold_rows"] == 50
        and validation["readable_rows"] == 50
        and not validation["secret_hits"]
    )
    if not required_ok:
        print("validation이 완전히 통과하지 않아 cleanup을 실행하지 않습니다.", file=sys.stderr)
        return []
    deleted: list[str] = []
    for path in candidates:
        if path.exists() and path.is_file():
            path.unlink()
            deleted.append(str(path))
    return deleted


def print_dry_run(result: GenerationResult) -> None:
    """dry-run 요약을 출력한다."""

    print("DRY RUN: 파일을 생성하지 않습니다.")
    print("record_count", len(result.records))
    print("unique_id_count", len({record["id"] for record in result.records}))
    print("task_counts", dict(Counter(record["task_family"] for record in result.records)))
    print("status_counts", dict(Counter(row["gold_generation_status"] for row in result.validation_rows)))
    print("warning_count", sum(row["warning_count"] for row in result.validation_rows))


def load_jsonl_records(path: Path) -> list[dict[str, Any]]:
    """기존 JSONL record를 읽는다."""

    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                records.append(json.loads(line))
    return records


def write_resolution_outputs(
    args: argparse.Namespace,
    records: list[dict[str, Any]],
    summary: dict[str, Any],
) -> dict[str, Any]:
    """warning resolution 결과를 모든 최종 산출물에 반영한다."""

    readable_rows, validation_rows = rows_from_records(records)
    write_jsonl(args.output_jsonl, records)
    write_csv(args.output_readable_csv, readable_rows)
    write_csv(args.output_validation_csv, validation_rows)
    write_excel(args.output_xlsx, readable_rows, validation_rows)
    update_approved_csv_for_resolution(args.output_approved_csv, records)
    validation = validate_outputs(
        args,
        GenerationResult(records, readable_rows, validation_rows, pd.DataFrame()),
    )
    write_warning_resolution_report(
        args.output_warning_resolution_report,
        summary,
        records,
        validation_rows,
    )
    append_resolution_to_generation_report(
        args.output_report,
        summary,
        Counter(row["gold_generation_status"] for row in validation_rows),
    )
    update_guide_for_resolution(args.output_guide)
    return validation


def run_warning_resolution(args: argparse.Namespace) -> int:
    """기존 최종 정답지의 warning을 보완한다."""

    if not args.output_jsonl.exists():
        raise FileNotFoundError(f"기존 JSONL이 없습니다: {args.output_jsonl}")
    records = load_jsonl_records(args.output_jsonl)
    if len(records) != 50:
        raise ValueError(f"JSONL record 수는 50이어야 합니다. 현재 {len(records)}")
    original_ids = [record["id"] for record in records]
    summary = resolve_warning_records(records, args.chunks_jsonl, args.eval_dir)
    if [record["id"] for record in records] != original_ids:
        raise ValueError("warning resolution 중 record 순서 또는 id 구성이 변경되었습니다.")
    readable_rows, validation_rows = rows_from_records(records)
    status_counts = Counter(row["gold_generation_status"] for row in validation_rows)
    resolution_counts = Counter(
        record.get("warning_resolution_status")
        for record in records
        if record.get("warning_resolution_status")
    )
    if args.dry_run:
        print("DRY RUN: warning resolution 결과를 파일에 쓰지 않습니다.")
        print("old_warning_count", len(summary["old_warning_ids"]))
        print("resolved", summary["resolved"])
        print("accepted_warning", summary["accepted_warning"])
        print("unresolved_needs_fix", summary["unresolved_needs_fix"])
        print("status_counts", dict(status_counts))
        print("resolution_counts", dict(resolution_counts))
        return 0
    validation = write_resolution_outputs(args, records, summary)
    print(json.dumps(to_builtin(validation), ensure_ascii=False, indent=2))
    print("resolution_counts", dict(resolution_counts))
    return 0 if not summary["unresolved_needs_fix"] else 4


def main() -> int:
    """스크립트 진입점이다."""

    args = parse_args()
    if args.resolve_warnings:
        return run_warning_resolution(args)

    result = build_generation(args)
    if args.dry_run:
        print_dry_run(result)
        return 0

    write_jsonl(args.output_jsonl, result.records)
    write_csv(args.output_readable_csv, result.readable_rows)
    write_csv(args.output_validation_csv, result.validation_rows)
    result.approved_df.to_csv(args.output_approved_csv, index=False, encoding="utf-8-sig")
    write_excel(args.output_xlsx, result.readable_rows, result.validation_rows)
    write_guide(args.output_guide, len(result.records))
    update_review_workbook(args.review_xlsx, result.approved_df)
    validation = validate_outputs(args, result)
    candidates = cleanup_candidates(args)
    cleanup_deleted = run_cleanup(args, validation, result)
    write_report(args.output_report, result, cleanup_deleted, candidates)
    write_cleanup_manifest(
        args.output_cleanup_manifest,
        args,
        validation,
        cleanup_deleted,
        candidates,
        args.cleanup,
    )
    print(json.dumps(to_builtin(validation), ensure_ascii=False, indent=2))
    if validation["jsonl_count"] != 50 or validation["unique_id_count"] != 50:
        return 2
    if validation["secret_hits"]:
        return 3
    print(f"completed_at {datetime.now().isoformat(timespec='seconds')}")
    print(f"python {sys.version.split()[0]} platform {platform.platform()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
